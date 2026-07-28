"""Pure-logic tests: designer schema, parameters, filters, exports,
recurrence, retention, and distribution validation.

None of these needs a database. They are deliberately side-effect-free,
which is exactly what makes them testable at this level.
"""

from __future__ import annotations

import io
import json
import uuid
import xml.etree.ElementTree as ET
import zipfile
from datetime import UTC, date, datetime, timedelta
from typing import Any
from zoneinfo import ZoneInfo

import pytest
from openpyxl import load_workbook
from shared_core.exceptions.conflict import ConflictError
from shared_core.exceptions.validation import ValidationError

from app.archive.retention import (
    ensure_purgeable,
    is_expired,
    retention_deadline,
    verify_integrity,
)
from app.distribution.channels import new_share_token, sign_webhook
from app.export.engine import (
    FORMAT_SPECS,
    export,
    export_pdf_protected,
    safe_filename,
)
from app.export.text_formats import theme_colors
from app.filters.engine import FilterClause, apply_filters, matches, parse_clauses
from app.models.enums import (
    ArchiveStatus,
    ChartKind,
    DistributionChannel,
    ExportFormat,
    FilterOperator,
    ParameterKind,
    ScheduleFrequency,
    SectionKind,
)
from app.models.report_archive import ReportArchive
from app.models.report_parameter import ReportParameter
from app.parameters.binding import BoundParameters, bind, coerce
from app.renderer.document import RenderedColumn, RenderedReport, RenderedSection
from app.renderer.engine import _aggregate, _chart_series, render_text
from app.reports.designer.schema import Branding, parse_definition
from app.scheduler.recurrence import compute_next_run, validate_schedule
from app.services.template import bump_minor
from app.validators.distribution import validate_target

ROWS: list[dict[str, Any]] = [
    {"name": "db-1", "env": "prod", "cpu": 91.5, "seen": "2026-07-20T10:00:00Z", "owner": None},
    {"name": "db-2", "env": "dev", "cpu": 12.0, "seen": "2026-07-27T10:00:00Z", "owner": "ops"},
    {"name": "web-1", "env": "prod", "cpu": 55.0, "seen": "2026-07-25T10:00:00Z", "owner": "web"},
]


def clause(field: str, operator: FilterOperator, value: Any = None) -> FilterClause:
    return FilterClause(field=field, operator=operator, value=value)


def names(rows: list[dict[str, Any]]) -> list[str]:
    return [row["name"] for row in rows]


class TestDesignerSchema:
    def test_a_valid_definition_parses(self) -> None:
        definition = parse_definition(
            {
                "title": "Fleet",
                "sections": [
                    {"key": "intro", "kind": "text", "text": "Hello"},
                    {
                        "key": "hosts",
                        "kind": "table",
                        "query": {"source": "inventory", "path": "/inventory/assets"},
                        "columns": [{"key": "name", "label": "Host"}],
                    },
                ],
            }
        )
        assert [section.key for section in definition.sections] == ["intro", "hosts"]

    def test_referenced_sources_are_deduplicated_in_order(self) -> None:
        definition = parse_definition(
            {
                "title": "T",
                "sections": [
                    {
                        "key": "a",
                        "kind": "table",
                        "query": {"source": "inventory", "path": "/x"},
                        "columns": [{"key": "k", "label": "K"}],
                    },
                    {
                        "key": "b",
                        "kind": "table",
                        "query": {"source": "inventory", "path": "/y"},
                        "columns": [{"key": "k", "label": "K"}],
                    },
                    {
                        "key": "c",
                        "kind": "table",
                        "query": {"source": "alerting", "path": "/alerts"},
                        "columns": [{"key": "k", "label": "K"}],
                    },
                ],
            }
        )
        assert [str(source) for source in definition.referenced_sources()] == [
            "inventory",
            "alerting",
        ]

    @pytest.mark.parametrize(
        ("section", "problem"),
        [
            ({"key": "t", "kind": "table"}, "requires a query"),
            (
                {
                    "key": "t",
                    "kind": "table",
                    "query": {"source": "inventory", "path": "/x"},
                },
                "requires columns",
            ),
            ({"key": "c", "kind": "chart"}, "requires a query"),
            (
                {"key": "c", "kind": "chart", "query": {"source": "inventory", "path": "/x"}},
                "requires a chart spec",
            ),
            ({"key": "m", "kind": "metric"}, "requires a query"),
            (
                {
                    "key": "m",
                    "kind": "metric",
                    "metric_aggregate": "avg",
                    "query": {"source": "inventory", "path": "/x"},
                },
                "requires a metric_key",
            ),
            ({"key": "h", "kind": "heading"}, "requires text"),
        ],
    )
    def test_a_section_that_cannot_render_is_rejected(
        self, section: dict[str, Any], problem: str
    ) -> None:
        """Catching this at authoring time is the whole value.

        The alternative is a scheduled report failing at 03:00 because a
        table section was saved with no query.
        """
        with pytest.raises(Exception, match=problem):
            parse_definition({"title": "T", "sections": [section]})

    def test_duplicate_section_keys_are_rejected(self) -> None:
        with pytest.raises(Exception, match="Duplicate section keys"):
            parse_definition(
                {
                    "title": "T",
                    "sections": [
                        {"key": "a", "kind": "text", "text": "x"},
                        {"key": "a", "kind": "text", "text": "y"},
                    ],
                }
            )

    def test_branding_defaults_are_sensible(self) -> None:
        definition = parse_definition({"title": "T", "sections": []})
        assert definition.branding.company_name == "AI-IOS"
        assert definition.branding.show_page_numbers is True


class TestParameterBinding:
    def _declared(self, **overrides: Any) -> ReportParameter:
        defaults: dict[str, Any] = {
            "organization_id": uuid.uuid4(),
            "template_id": uuid.uuid4(),
            "key": "environment",
            "label": "Environment",
            "kind": ParameterKind.STRING,
            "required": False,
            "default_value": None,
            "allowed_values": [],
            "display_order": 0,
        }
        return ReportParameter(**{**defaults, **overrides})

    @pytest.mark.parametrize(
        ("kind", "raw", "expected"),
        [
            (ParameterKind.STRING, 7, "7"),
            (ParameterKind.INTEGER, "7", 7),
            (ParameterKind.INTEGER, 7.0, 7),
            (ParameterKind.NUMBER, "7.5", 7.5),
            (ParameterKind.BOOLEAN, "yes", True),
            (ParameterKind.BOOLEAN, "off", False),
            (ParameterKind.DATE, "2026-07-28", date(2026, 7, 28)),
            (ParameterKind.DATETIME, "2026-07-28T10:00:00Z", datetime(2026, 7, 28, 10, tzinfo=UTC)),
        ],
    )
    def test_representation_is_coerced(self, kind: ParameterKind, raw: Any, expected: Any) -> None:
        assert coerce("k", kind, raw) == expected

    @pytest.mark.parametrize(
        ("kind", "raw"),
        [
            (ParameterKind.INTEGER, 7.5),
            (ParameterKind.INTEGER, True),
            (ParameterKind.INTEGER, "seven"),
            (ParameterKind.NUMBER, True),
            (ParameterKind.BOOLEAN, "maybe"),
            (ParameterKind.DATE, "not-a-date"),
            (ParameterKind.DATETIME, "not-a-datetime"),
            (ParameterKind.UUID, "not-a-uuid"),
        ],
    )
    def test_information_loss_is_rejected(self, kind: ParameterKind, raw: Any) -> None:
        """Silently truncating a typed value is how wrong numbers reach a report."""
        with pytest.raises(ValidationError):
            coerce("k", kind, raw)

    def test_none_stays_none(self) -> None:
        assert coerce("k", ParameterKind.INTEGER, None) is None

    def test_defaults_are_applied(self) -> None:
        bound = bind([self._declared(default_value="prod")], {})
        assert bound.get("environment") == "prod"

    def test_supplied_beats_default(self) -> None:
        bound = bind([self._declared(default_value="prod")], {"environment": "dev"})
        assert bound.get("environment") == "dev"

    def test_missing_required_is_rejected(self) -> None:
        with pytest.raises(ValidationError, match="Missing required"):
            bind([self._declared(required=True)], {})

    def test_unknown_parameter_is_rejected(self) -> None:
        """A typo'd name that silently does nothing produces a report that
        looks right but was filtered differently than the user asked for.
        """
        with pytest.raises(ValidationError, match="Unknown report parameter"):
            bind([self._declared()], {"enviroment": "prod"})

    def test_allowed_values_are_enforced(self) -> None:
        declared = self._declared(allowed_values=["prod", "dev"])
        with pytest.raises(ValidationError, match="must be one of"):
            bind([declared], {"environment": "staging"})

    def test_query_params_render_every_type(self) -> None:
        bound = BoundParameters(
            values={
                "text": "prod",
                "when": datetime(2026, 7, 28, tzinfo=UTC),
                "day": date(2026, 7, 28),
                "flag": True,
                "many": ["a", "b"],
                "absent": None,
            }
        )
        rendered = bound.as_query_params()
        assert rendered["text"] == "prod"
        assert rendered["when"].startswith("2026-07-28T00:00:00")
        assert rendered["day"] == "2026-07-28"
        assert rendered["flag"] == "true"
        assert rendered["many"] == "a,b"
        assert "absent" not in rendered


class TestFilters:
    @pytest.mark.parametrize(
        ("field", "operator", "value", "expected"),
        [
            ("env", FilterOperator.EQUALS, "prod", ["db-1", "web-1"]),
            ("env", FilterOperator.NOT_EQUALS, "prod", ["db-2"]),
            ("cpu", FilterOperator.GREATER_THAN, 50, ["db-1", "web-1"]),
            ("cpu", FilterOperator.GREATER_OR_EQUAL, 55, ["db-1", "web-1"]),
            ("cpu", FilterOperator.LESS_THAN, 55, ["db-2"]),
            ("cpu", FilterOperator.LESS_OR_EQUAL, 55, ["db-2", "web-1"]),
            ("env", FilterOperator.IN, ["dev"], ["db-2"]),
            ("env", FilterOperator.NOT_IN, ["prod"], ["db-2"]),
            ("name", FilterOperator.CONTAINS, "web", ["web-1"]),
            ("name", FilterOperator.STARTS_WITH, "db", ["db-1", "db-2"]),
            ("cpu", FilterOperator.BETWEEN, [20, 60], ["web-1"]),
            ("owner", FilterOperator.IS_NULL, None, ["db-1"]),
            ("owner", FilterOperator.IS_NOT_NULL, None, ["db-2", "web-1"]),
        ],
    )
    def test_every_operator(
        self, field: str, operator: FilterOperator, value: Any, expected: list[str]
    ) -> None:
        assert names(apply_filters(ROWS, [clause(field, operator, value)])) == expected

    def test_iso_strings_compare_against_real_datetimes(self) -> None:
        """Sources return ISO strings; a bound parameter is a real object."""
        cutoff = datetime(2026, 7, 25, tzinfo=UTC)
        assert names(
            apply_filters(ROWS, [clause("seen", FilterOperator.GREATER_OR_EQUAL, cutoff)])
        ) == ["db-2", "web-1"]

    def test_clauses_combine_with_and(self) -> None:
        assert names(
            apply_filters(
                ROWS,
                [
                    clause("env", FilterOperator.EQUALS, "prod"),
                    clause("cpu", FilterOperator.LESS_THAN, 60),
                ],
            )
        ) == ["web-1"]

    def test_no_clauses_passes_everything_through(self) -> None:
        assert apply_filters(ROWS, []) == ROWS

    def test_nested_paths_resolve(self) -> None:
        rows = [{"name": "a", "meta": {"owner": {"team": "ops"}}}, {"name": "b", "meta": {}}]
        assert names(
            apply_filters(rows, [clause("meta.owner.team", FilterOperator.EQUALS, "ops")])
        ) == ["a"]

    def test_a_missing_field_never_raises(self) -> None:
        """One absent optional field must not fail an entire report."""
        assert apply_filters(ROWS, [clause("nope.deep", FilterOperator.EQUALS, "x")]) == []

    def test_incomparable_values_simply_do_not_match(self) -> None:
        rows = [{"name": "a", "v": "not-a-number"}]
        assert apply_filters(rows, [clause("v", FilterOperator.GREATER_THAN, 5)]) == []

    def test_between_with_incomparable_bounds(self) -> None:
        rows = [{"name": "a", "v": "text"}]
        assert apply_filters(rows, [clause("v", FilterOperator.BETWEEN, [1, 2])]) == []

    def test_parse_clauses_round_trips(self) -> None:
        parsed = parse_clauses([{"field": "env", "operator": "eq", "value": "prod"}])
        assert parsed[0].field == "env"
        assert parsed[0].operator is FilterOperator.EQUALS

    @pytest.mark.parametrize(
        ("raw", "problem"),
        [
            ({}, "requires a 'field'"),
            ({"field": "x", "operator": "nope"}, "unknown operator"),
            ({"field": "x", "operator": "in", "value": "notalist"}, "requires a list"),
            ({"field": "x", "operator": "between", "value": [1]}, "exactly two values"),
            ({"field": "x", "operator": "eq"}, "requires a value"),
        ],
    )
    def test_malformed_clauses_are_rejected(self, raw: dict[str, Any], problem: str) -> None:
        with pytest.raises(ValidationError, match=problem):
            parse_clauses([raw])

    def test_null_only_operators_need_no_value(self) -> None:
        parsed = parse_clauses([{"field": "owner", "operator": "is_null"}])
        assert matches({"owner": None}, parsed[0])


class TestAggregatesAndCharts:
    @pytest.mark.parametrize(
        ("how", "expected"),
        [("count", 3.0), ("sum", 4.0), ("avg", 2.0), ("min", 1.0), ("max", 3.0)],
    )
    def test_every_aggregate(self, how: str, expected: float) -> None:
        rows: list[dict[str, Any]] = [{"v": 1}, {"v": 3}, {"v": "n/a"}]
        assert _aggregate(rows, "v", how) == expected

    def test_booleans_are_not_treated_as_numbers(self) -> None:
        """``True`` is not the number 1 in a metric column."""
        assert _aggregate([{"v": True}, {"v": 4}], "v", "avg") == 4.0

    def test_no_numeric_values_is_zero_not_an_error(self) -> None:
        assert _aggregate([{"v": "x"}], "v", "sum") == 0.0

    def test_a_metric_without_a_key_is_zero(self) -> None:
        assert _aggregate(ROWS, None, "sum") == 0.0

    def test_chart_totals_are_preserved_when_truncated(self) -> None:
        """A truncated chart that under-reports is a correctness bug."""
        rows = [{"l": f"c{index}", "v": 1} for index in range(10)]
        series = _chart_series(rows, "l", "v", 3)
        assert len(series) == 3
        assert sum(value for _label, value in series) == 10.0
        assert series[-1][0] == "Other"

    def test_chart_groups_and_orders_by_magnitude(self) -> None:
        series = _chart_series(ROWS, "env", "cpu", 12)
        assert series[0] == ("prod", 146.5)
        assert series[1] == ("dev", 12.0)

    def test_missing_labels_become_none_bucket(self) -> None:
        series = _chart_series([{"v": 1}], "l", "v", 5)
        assert series == [("(none)", 1.0)]

    def test_non_numeric_chart_values_count_as_one(self) -> None:
        series = _chart_series([{"l": "a", "v": "x"}, {"l": "a", "v": "y"}], "l", "v", 5)
        assert series == [("a", 2.0)]


class TestSectionTextRendering:
    def test_parameters_interpolate(self) -> None:
        assert render_text("Env {{ environment }}", {"environment": "prod"}) == "Env prod"

    def test_an_undefined_variable_raises(self) -> None:
        """A sentence that silently loses its subject is worse than an error."""
        with pytest.raises(ValidationError, match="failed to render"):
            render_text("Env {{ missing }}", {})

    def test_the_environment_is_sandboxed(self) -> None:
        """A template is user-authored content, so it must not reach internals."""
        with pytest.raises(ValidationError):
            render_text("{{ ''.__class__.__mro__ }}", {})


def _sample_report() -> RenderedReport:
    return RenderedReport(
        title="Fleet / Q3 <2026>",
        subtitle='A "quarterly" review & outlook',
        branding=Branding(company_name="AI-IOS", theme="ocean", footer_text="Confidential"),
        generated_at=datetime(2026, 7, 28, 12, tzinfo=UTC),
        parameters={"environment": "prod"},
        sections=[
            RenderedSection(key="t", kind=SectionKind.TEXT, title="Notes", text="Line."),
            RenderedSection(
                key="m",
                kind=SectionKind.METRIC,
                title="Total",
                metric_value=412.0,
                metric_label="Assets",
            ),
            RenderedSection(
                key="tb",
                kind=SectionKind.TABLE,
                title="Hosts",
                columns=[RenderedColumn("name", "Host"), RenderedColumn("cpu", "CPU %")],
                rows=[{"name": "db-1</td>", "cpu": 91.5}, {"name": 'q"q', "cpu": None}],
            ),
            RenderedSection(
                key="c",
                kind=SectionKind.CHART,
                title="By env",
                chart_kind=ChartKind.BAR,
                chart_points=[("prod", 10.0), ("dev", 4.0)],
            ),
            RenderedSection(key="e", kind=SectionKind.TABLE, title="Down", error="HTTP 503"),
        ],
    )


class TestExporters:
    @pytest.mark.parametrize("export_format", list(ExportFormat))
    def test_every_format_produces_content(self, export_format: ExportFormat) -> None:
        artifact = export(_sample_report(), export_format)
        assert artifact.size_bytes > 0
        assert artifact.content_type == FORMAT_SPECS[export_format].content_type
        assert artifact.checksum_sha256

    def test_json_parses_and_carries_every_section(self) -> None:
        payload = json.loads(export(_sample_report(), ExportFormat.JSON).content)
        assert payload["title"] == "Fleet / Q3 <2026>"
        assert len(payload["sections"]) == 5

    def test_xml_parses(self) -> None:
        root = ET.fromstring(export(_sample_report(), ExportFormat.XML).content)
        assert root.tag == "report"
        sections = root.find("sections")
        assert sections is not None
        assert len(sections) == 5

    def test_xlsx_is_a_real_workbook_with_one_sheet_per_section(self) -> None:
        content = export(_sample_report(), ExportFormat.XLSX).content
        assert zipfile.is_zipfile(io.BytesIO(content))
        workbook = load_workbook(io.BytesIO(content))
        assert workbook.sheetnames[0] == "Summary"
        assert "Hosts" in workbook.sheetnames

    def test_xlsx_keeps_numbers_native(self) -> None:
        """A spreadsheet where 9 sorts after 10 is useless for analysis."""
        workbook = load_workbook(io.BytesIO(export(_sample_report(), ExportFormat.XLSX).content))
        sheet = workbook["Hosts"]
        assert sheet.cell(row=2, column=2).value == 91.5

    def test_pdf_is_a_real_document(self) -> None:
        content = export(_sample_report(), ExportFormat.PDF).content
        assert content.startswith(b"%PDF")

    def test_pdf_password_really_encrypts(self) -> None:
        content = export_pdf_protected(_sample_report(), password="s3cret").content
        assert b"/Encrypt" in content

    def test_pdf_signature_block_is_added(self) -> None:
        signed = export_pdf_protected(_sample_report(), signed_by="ops@example.com").content
        plain = export(_sample_report(), ExportFormat.PDF).content
        assert len(signed) > len(plain)

    def test_csv_escapes_quotes(self) -> None:
        text = export(_sample_report(), ExportFormat.CSV).content.decode()
        assert '"q""q"' in text

    def test_html_escapes_markup(self) -> None:
        """A row containing markup must render as text, not execute."""
        text = export(_sample_report(), ExportFormat.HTML).content.decode()
        assert "&lt;/td&gt;" in text
        assert "<td>db-1</td>," not in text

    def test_markdown_escapes_pipes(self) -> None:
        report = _sample_report()
        report.sections[2].rows = [{"name": "a|b", "cpu": 1}]
        text = export(report, ExportFormat.MARKDOWN).content.decode()
        assert r"a\|b" in text

    def test_failed_sections_are_visible_in_every_text_format(self) -> None:
        """Hiding the gap would be worse than showing it."""
        for export_format in (ExportFormat.MARKDOWN, ExportFormat.HTML, ExportFormat.XML):
            text = export(_sample_report(), export_format).content.decode()
            assert "503" in text

    def test_csv_without_tables_still_produces_a_document(self) -> None:
        report = _sample_report()
        report.sections = [RenderedSection(key="t", kind=SectionKind.TEXT, text="only text")]
        assert export(report, ExportFormat.CSV).size_bytes > 0

    def test_multiple_tables_are_segmented(self) -> None:
        report = _sample_report()
        report.sections.append(
            RenderedSection(
                key="tb2",
                kind=SectionKind.TABLE,
                title="Second",
                columns=[RenderedColumn("x", "X")],
                rows=[{"x": 1}],
            )
        )
        text = export(report, ExportFormat.CSV).content.decode()
        assert "# Hosts" in text
        assert "# Second" in text

    @pytest.mark.parametrize(
        ("title", "expected"),
        [
            ("Fleet / Q3", "Fleet-Q3.csv"),
            ('bad"name\nhere', "bad-name-here.csv"),
            ("...", "report.csv"),
        ],
    )
    def test_filenames_are_header_safe(self, title: str, expected: str) -> None:
        """An unescaped newline in Content-Disposition is header injection."""
        assert safe_filename(title, ExportFormat.CSV) == expected

    def test_unknown_theme_falls_back_rather_than_failing(self) -> None:
        assert theme_colors("nonexistent") == theme_colors("slate")

    def test_report_totals(self) -> None:
        report = _sample_report()
        assert report.total_rows == 2
        assert [section.key for section in report.failed_sections] == ["e"]


class TestRecurrence:
    START = datetime(2026, 1, 31, 9, 0, tzinfo=UTC)
    NOW = datetime(2026, 7, 28, 12, 0, tzinfo=UTC)

    def test_a_past_one_time_schedule_is_finished(self) -> None:
        """Returning the past would make the worker re-run it forever."""
        assert (
            compute_next_run(ScheduleFrequency.ONE_TIME, starts_at=self.START, after=self.NOW)
            is None
        )

    def test_a_future_one_time_schedule_returns_its_start(self) -> None:
        future = self.NOW + timedelta(days=1)
        assert (
            compute_next_run(ScheduleFrequency.ONE_TIME, starts_at=future, after=self.NOW) == future
        )

    @pytest.mark.parametrize(
        ("frequency", "expected"),
        [
            (ScheduleFrequency.HOURLY, datetime(2026, 7, 28, 13, 0, tzinfo=UTC)),
            (ScheduleFrequency.DAILY, datetime(2026, 7, 29, 9, 0, tzinfo=UTC)),
            (ScheduleFrequency.WEEKLY, datetime(2026, 8, 1, 9, 0, tzinfo=UTC)),
        ],
    )
    def test_fixed_cadences(self, frequency: ScheduleFrequency, expected: datetime) -> None:
        assert compute_next_run(frequency, starts_at=self.START, after=self.NOW) == expected

    def test_monthly_clamps_to_the_last_valid_day(self) -> None:
        """Jan 31 + one month is Feb 28, not March 3.

        Rolling over would silently move a month-end compliance report
        into the following month.
        """
        assert compute_next_run(
            ScheduleFrequency.MONTHLY,
            starts_at=self.START,
            after=datetime(2026, 2, 1, tzinfo=UTC),
        ) == datetime(2026, 2, 28, 9, 0, tzinfo=UTC)

    def test_cron_is_delegated_to_shared_core(self) -> None:
        assert compute_next_run(
            ScheduleFrequency.CRON,
            cron_expression="0 6 * * *",
            starts_at=self.START,
            after=self.NOW,
        ) == datetime(2026, 7, 29, 6, 0, tzinfo=UTC)

    def test_local_wall_clock_survives_a_dst_transition(self) -> None:
        """A 09:00 Berlin report must stay at 09:00 Berlin across the change."""
        berlin_start = datetime(2026, 3, 20, 8, 0, tzinfo=UTC)  # 09:00 CET
        result = compute_next_run(
            ScheduleFrequency.DAILY,
            timezone_name="Europe/Berlin",
            starts_at=berlin_start,
            after=datetime(2026, 3, 30, 12, 0, tzinfo=UTC),
        )
        assert result is not None
        assert result.astimezone(ZoneInfo("Europe/Berlin")).hour == 9

    def test_a_naive_start_is_treated_as_utc(self) -> None:
        result = compute_next_run(
            ScheduleFrequency.DAILY,
            starts_at=datetime(2026, 1, 31, 9, 0),
            after=self.NOW,
        )
        assert result is not None
        assert result.tzinfo is not None

    @pytest.mark.parametrize(
        ("frequency", "cron", "timezone_name", "problem"),
        [
            (ScheduleFrequency.CRON, None, "UTC", "requires a cron_expression"),
            (ScheduleFrequency.CRON, "not a cron", "UTC", "Invalid cron"),
            (ScheduleFrequency.DAILY, None, "Mars/Olympus", "Unknown time zone"),
            (ScheduleFrequency.DAILY, "0 6 * * *", "UTC", "must not carry a cron_expression"),
        ],
    )
    def test_invalid_configurations_are_rejected(
        self,
        frequency: ScheduleFrequency,
        cron: str | None,
        timezone_name: str,
        problem: str,
    ) -> None:
        with pytest.raises(ValidationError, match=problem):
            validate_schedule(frequency, cron, timezone_name)


class TestRetention:
    def _archive(self, **overrides: Any) -> ReportArchive:
        content = b"archived-bytes"
        defaults: dict[str, Any] = {
            "organization_id": uuid.uuid4(),
            "title": "Report",
            "export_format": ExportFormat.CSV,
            "filename": "r.csv",
            "content_type": "text/csv",
            "size_bytes": len(content),
            "checksum_sha256": __import__("hashlib").sha256(content).hexdigest(),
            "version": 1,
            "status": ArchiveStatus.ACTIVE,
            "archived_at": datetime(2026, 1, 1, tzinfo=UTC),
            "retention_until": datetime(2026, 6, 1, tzinfo=UTC),
            "content": content,
        }
        return ReportArchive(**{**defaults, **overrides})

    def test_deadline_is_computed_from_the_archive_moment(self) -> None:
        assert retention_deadline(
            archived_at=datetime(2026, 1, 1, tzinfo=UTC), retention_days=30
        ) == datetime(2026, 1, 31, tzinfo=UTC)

    def test_integrity_passes_for_untouched_content(self) -> None:
        assert verify_integrity(self._archive())

    def test_integrity_fails_when_content_changed(self) -> None:
        """Not that bytes cannot change -- that a change cannot go unnoticed."""
        assert not verify_integrity(self._archive(content=b"tampered"))

    def test_expiry(self) -> None:
        archive = self._archive()
        assert is_expired(archive, moment=datetime(2026, 7, 1, tzinfo=UTC))
        assert not is_expired(archive, moment=datetime(2026, 5, 1, tzinfo=UTC))

    def test_indefinite_retention_never_expires(self) -> None:
        """ "Keep forever" must not be mistaken for "expired immediately"."""
        assert not is_expired(self._archive(retention_until=None))

    def test_purging_inside_the_retention_window_is_refused(self) -> None:
        with pytest.raises(ConflictError, match="retained until"):
            ensure_purgeable(self._archive(), moment=datetime(2026, 5, 1, tzinfo=UTC))

    def test_purging_after_retention_is_allowed(self) -> None:
        ensure_purgeable(self._archive(), moment=datetime(2026, 7, 1, tzinfo=UTC))

    def test_purging_twice_is_refused(self) -> None:
        with pytest.raises(ConflictError, match="already been purged"):
            ensure_purgeable(
                self._archive(status=ArchiveStatus.PURGED),
                moment=datetime(2026, 7, 1, tzinfo=UTC),
            )

    def test_a_string_status_from_the_database_is_handled(self) -> None:
        """A row read back from Postgres yields a raw ``str``."""
        archive = self._archive()
        archive.status = "purged"  # type: ignore[assignment]
        with pytest.raises(ConflictError, match="already been purged"):
            ensure_purgeable(archive, moment=datetime(2026, 7, 1, tzinfo=UTC))


class TestDistributionPrimitives:
    def test_share_tokens_are_unguessable_and_unique(self) -> None:
        tokens = {new_share_token() for _ in range(100)}
        assert len(tokens) == 100
        assert all(len(token) >= 40 for token in tokens)

    def test_webhook_signature_is_stable_and_covers_the_timestamp(self) -> None:
        first = sign_webhook("secret", "2026-07-28T00:00:00Z", b"payload")
        assert first == sign_webhook("secret", "2026-07-28T00:00:00Z", b"payload")
        assert first != sign_webhook("secret", "2026-07-28T00:00:01Z", b"payload")
        assert first != sign_webhook("other", "2026-07-28T00:00:00Z", b"payload")
        assert first.startswith("sha256=")

    @pytest.mark.parametrize(
        ("channel", "target"),
        [
            (DistributionChannel.EMAIL, "ops@example.com"),
            (DistributionChannel.WEBHOOK, "https://hooks.example.com/report"),
            (DistributionChannel.DOWNLOAD, "anything"),
        ],
    )
    def test_valid_targets_pass(self, channel: DistributionChannel, target: str) -> None:
        validate_target(channel, target)

    @pytest.mark.parametrize(
        ("channel", "target", "problem"),
        [
            (DistributionChannel.EMAIL, "not-an-email", "not a valid email"),
            (DistributionChannel.WEBHOOK, "/relative/path", "absolute http"),
            (DistributionChannel.WEBHOOK, "file:///etc/passwd", "absolute http"),
            (DistributionChannel.EMAIL, "   ", "requires a target"),
        ],
    )
    def test_invalid_targets_are_rejected_at_registration(
        self, channel: DistributionChannel, target: str, problem: str
    ) -> None:
        """A typo'd webhook URL must fail now, not at 03:00."""
        with pytest.raises(ValidationError, match=problem):
            validate_target(channel, target)

    def test_over_long_targets_are_rejected(self) -> None:
        with pytest.raises(ValidationError, match="exceeds"):
            validate_target(DistributionChannel.DOWNLOAD, "x" * 2000)


class TestVersioning:
    @pytest.mark.parametrize(
        ("current", "expected"),
        [("1.0.0", "1.1.0"), ("2.7.3", "2.8.0"), ("weird", "weird.1"), ("1.0", "1.0.1")],
    )
    def test_minor_version_bumping(self, current: str, expected: str) -> None:
        assert bump_minor(current) == expected
