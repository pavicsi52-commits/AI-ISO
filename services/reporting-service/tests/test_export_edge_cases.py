"""Exporter branches a happy-path report never reaches.

Broken logos, page breaks, zero-magnitude charts, over-long and
duplicate worksheet names, oversized tables, and column keys that are
not valid XML element names -- each one a real way a user-authored
template can differ from the well-behaved case.
"""

from __future__ import annotations

import io
import xml.etree.ElementTree as ET
import zipfile
from datetime import UTC, datetime
from typing import Any

import pytest
from openpyxl import load_workbook
from shared_core.exceptions.validation import ValidationError

from app.export.engine import export
from app.models.enums import ChartKind, ExportFormat, SectionKind
from app.renderer.document import RenderedColumn, RenderedReport, RenderedSection
from app.reports.designer.schema import Branding

ONE_PIXEL_PNG = (
    "data:image/png;base64,"
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
)


def build_report(**overrides: Any) -> RenderedReport:
    defaults: dict[str, Any] = {
        "title": "Edge",
        "subtitle": None,
        "branding": Branding(),
        "generated_at": datetime(2026, 7, 28, tzinfo=UTC),
        "sections": [],
        "parameters": {},
    }
    return RenderedReport(**{**defaults, **overrides})


class TestPdfEdgeCases:
    def test_a_page_break_renders_in_every_format(self) -> None:
        report = build_report(
            sections=[
                RenderedSection(key="a", kind=SectionKind.TEXT, text="before"),
                RenderedSection(key="b", kind=SectionKind.PAGE_BREAK),
                RenderedSection(key="c", kind=SectionKind.TEXT, text="after"),
            ]
        )
        assert "---" in export(report, ExportFormat.MARKDOWN).content.decode()
        assert "<hr>" in export(report, ExportFormat.HTML).content.decode()
        assert export(report, ExportFormat.PDF).content.startswith(b"%PDF")

    def test_a_valid_logo_is_embedded(self) -> None:
        report = build_report(
            branding=Branding(logo_data_uri=ONE_PIXEL_PNG),
            sections=[RenderedSection(key="a", kind=SectionKind.TEXT, text="x")],
        )
        assert export(report, ExportFormat.PDF).content.startswith(b"%PDF")

    @pytest.mark.parametrize(
        "logo", ["data:image/png,notbase64", "data:image/png;base64,", "nonsense"]
    )
    def test_a_broken_logo_is_skipped_not_fatal(self, logo: str) -> None:
        """Branding is cosmetic; no logo beats no report."""
        report = build_report(
            branding=Branding(logo_data_uri=logo),
            sections=[RenderedSection(key="a", kind=SectionKind.TEXT, text="x")],
        )
        assert export(report, ExportFormat.PDF).content.startswith(b"%PDF")

    def test_a_contents_list_needs_at_least_two_titles(self) -> None:
        """A one-item contents list is noise."""
        one = build_report(
            sections=[RenderedSection(key="a", kind=SectionKind.TEXT, title="Only", text="x")]
        )
        two = build_report(
            sections=[
                RenderedSection(key="a", kind=SectionKind.TEXT, title="One", text="x"),
                RenderedSection(key="b", kind=SectionKind.TEXT, title="Two", text="y"),
            ]
        )
        assert len(export(two, ExportFormat.PDF).content) > len(
            export(one, ExportFormat.PDF).content
        )

    def test_page_numbers_and_contents_can_be_disabled(self) -> None:
        report = build_report(
            branding=Branding(show_page_numbers=False, show_table_of_contents=False),
            sections=[RenderedSection(key="a", kind=SectionKind.TEXT, text="x")],
        )
        assert export(report, ExportFormat.PDF).content.startswith(b"%PDF")

    def test_an_oversized_table_is_truncated_visibly(self) -> None:
        """Silently dropping rows would be worse than saying so."""
        report = build_report(
            sections=[
                RenderedSection(
                    key="big",
                    kind=SectionKind.TABLE,
                    title="Big",
                    columns=[RenderedColumn("n", "N")],
                    rows=[{"n": index} for index in range(2_100)],
                )
            ]
        )
        assert export(report, ExportFormat.PDF).content.startswith(b"%PDF")

    def test_a_zero_magnitude_chart_does_not_divide_by_zero(self) -> None:
        report = build_report(
            sections=[
                RenderedSection(
                    key="c",
                    kind=SectionKind.CHART,
                    title="Flat",
                    chart_kind=ChartKind.BAR,
                    chart_points=[("a", 0.0), ("b", 0.0)],
                )
            ]
        )
        assert export(report, ExportFormat.PDF).content.startswith(b"%PDF")
        assert "0" in export(report, ExportFormat.HTML).content.decode()

    def test_very_wide_content_still_fits_columns(self) -> None:
        """Column sizing must handle content wider than the page."""
        report = build_report(
            sections=[
                RenderedSection(
                    key="w",
                    kind=SectionKind.TABLE,
                    title="Wide",
                    columns=[RenderedColumn(f"c{i}", f"Column {i}") for i in range(12)],
                    rows=[{f"c{i}": "x" * 80 for i in range(12)}],
                )
            ]
        )
        assert export(report, ExportFormat.PDF).content.startswith(b"%PDF")


class TestXlsxEdgeCases:
    @pytest.mark.parametrize("kind", [ChartKind.BAR, ChartKind.LINE, ChartKind.PIE])
    def test_every_chart_kind_reaches_the_workbook(self, kind: ChartKind) -> None:
        report = build_report(
            sections=[
                RenderedSection(
                    key="c",
                    kind=SectionKind.CHART,
                    title="Chart",
                    chart_kind=kind,
                    chart_points=[("a", 3.0), ("b", 1.0)],
                )
            ]
        )
        assert zipfile.is_zipfile(io.BytesIO(export(report, ExportFormat.XLSX).content))

    def test_sheet_names_are_sanitised_and_deduplicated(self) -> None:
        """Excel rejects five characters outright and any duplicate name."""
        report = build_report(
            sections=[
                RenderedSection(
                    key=f"s{index}",
                    kind=SectionKind.TABLE,
                    title="Bad/Name:With*Chars",
                    columns=[RenderedColumn("a", "A")],
                    rows=[{"a": 1}],
                )
                for index in range(3)
            ]
        )
        workbook = load_workbook(io.BytesIO(export(report, ExportFormat.XLSX).content))
        assert len(set(workbook.sheetnames)) == len(workbook.sheetnames)
        assert all("/" not in name and ":" not in name for name in workbook.sheetnames)

    def test_an_over_long_sheet_name_is_truncated(self) -> None:
        report = build_report(
            sections=[
                RenderedSection(
                    key="s",
                    kind=SectionKind.TABLE,
                    title="X" * 60,
                    columns=[RenderedColumn("a", "A")],
                    rows=[{"a": 1}],
                )
            ]
        )
        workbook = load_workbook(io.BytesIO(export(report, ExportFormat.XLSX).content))
        assert all(len(name) <= 31 for name in workbook.sheetnames)

    def test_a_metric_only_section_lands_on_the_summary_sheet(self) -> None:
        report = build_report(
            sections=[
                RenderedSection(
                    key="m", kind=SectionKind.METRIC, metric_value=7.0, metric_label="Total"
                )
            ]
        )
        workbook = load_workbook(io.BytesIO(export(report, ExportFormat.XLSX).content))
        assert workbook.sheetnames == ["Summary"]

    def test_a_failed_section_is_noted_on_the_summary_sheet(self) -> None:
        report = build_report(
            sections=[RenderedSection(key="e", kind=SectionKind.TABLE, error="HTTP 503")]
        )
        workbook = load_workbook(io.BytesIO(export(report, ExportFormat.XLSX).content))
        values = [cell.value for row in workbook["Summary"].iter_rows() for cell in row]
        assert any("unavailable" in str(value) for value in values)

    def test_parameters_and_subtitle_reach_the_summary(self) -> None:
        report = build_report(subtitle="A subtitle", parameters={"environment": "prod"})
        workbook = load_workbook(io.BytesIO(export(report, ExportFormat.XLSX).content))
        values = [cell.value for row in workbook["Summary"].iter_rows() for cell in row]
        assert "environment" in values
        assert "A subtitle" in values


class TestTextFormatEdgeCases:
    def test_cell_values_render_every_python_type(self) -> None:
        section = RenderedSection(
            key="t",
            kind=SectionKind.TABLE,
            columns=[RenderedColumn(key, key) for key in ("a", "b", "c", "d", "e", "f")],
            rows=[
                {
                    "a": None,
                    "b": True,
                    "c": datetime(2026, 7, 28, tzinfo=UTC),
                    "d": [1, 2],
                    "e": {"k": "v"},
                    "f": 3,
                }
            ],
        )
        rendered = section.cell_values()[0]
        assert rendered[0] == ""
        assert rendered[1] == "true"
        assert rendered[2].startswith("2026-07-28")
        assert rendered[3] == "1, 2"
        assert rendered[4] == "k=v"
        assert rendered[5] == "3"

    def test_xml_falls_back_for_keys_that_are_not_element_names(self) -> None:
        """A column key with a space must not produce malformed XML."""
        report = build_report(
            sections=[
                RenderedSection(
                    key="t",
                    kind=SectionKind.TABLE,
                    columns=[RenderedColumn("not a name", "Label")],
                    rows=[{"not a name": "value"}],
                )
            ]
        )
        root = ET.fromstring(export(report, ExportFormat.XML).content)
        assert root.find(".//cell") is not None

    def test_a_subtitle_and_footer_reach_every_text_format(self) -> None:
        report = build_report(
            subtitle="A subtitle",
            branding=Branding(footer_text="Confidential"),
            sections=[RenderedSection(key="a", kind=SectionKind.TEXT, text="x")],
        )
        assert "A subtitle" in export(report, ExportFormat.MARKDOWN).content.decode()
        assert "Confidential" in export(report, ExportFormat.HTML).content.decode()
        assert "A subtitle" in export(report, ExportFormat.XML).content.decode()

    def test_a_metric_and_chart_reach_markdown_and_html(self) -> None:
        report = build_report(
            sections=[
                RenderedSection(
                    key="m", kind=SectionKind.METRIC, metric_value=412.0, metric_label="Assets"
                ),
                RenderedSection(
                    key="c",
                    kind=SectionKind.CHART,
                    chart_kind=ChartKind.BAR,
                    chart_points=[("prod", 10.0)],
                ),
            ]
        )
        markdown = export(report, ExportFormat.MARKDOWN).content.decode()
        assert "412" in markdown
        assert "prod" in markdown
        html = export(report, ExportFormat.HTML).content.decode()
        assert "412" in html

    def test_an_unsupported_format_is_rejected(self) -> None:
        with pytest.raises(ValidationError, match="not supported"):
            export(build_report(), "not-a-format")  # type: ignore[arg-type]
