"""Excel (XLSX) export via ``openpyxl``.

One worksheet per section rather than one flat sheet, because sections
legitimately have different column sets and stacking them would produce
a spreadsheet nobody can filter or pivot. A summary sheet leads, so the
workbook opens on something readable rather than on whichever section
happened to come first.
"""

from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.export.text_formats import theme_colors
from app.models.enums import ChartKind
from app.renderer.document import RenderedReport, RenderedSection

_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
_MAX_SHEET_NAME = 31
"""Excel's own hard limit on worksheet name length."""

_MAX_COLUMN_WIDTH = 60
_SAMPLE_ROWS = 200


def _safe_sheet_name(raw: str, used: set[str]) -> str:
    """Make a worksheet name Excel will actually accept.

    Excel rejects five characters outright and truncates past 31, and a
    duplicate name raises. Rather than let any of those surface as a
    failed export, the name is sanitised and de-duplicated with a
    numeric suffix.
    """
    cleaned = _INVALID_SHEET_CHARS.sub("-", raw).strip() or "Section"
    cleaned = cleaned[:_MAX_SHEET_NAME]
    candidate = cleaned
    suffix = 2
    while candidate.lower() in used:
        tail = f" ({suffix})"
        candidate = cleaned[: _MAX_SHEET_NAME - len(tail)] + tail
        suffix += 1
    used.add(candidate.lower())
    return candidate


def _style_header(sheet: Worksheet, columns: int, header_color: str) -> None:
    """Apply the theme's header styling and freeze the header row."""
    fill = PatternFill("solid", fgColor=header_color.lstrip("#"))
    font = Font(bold=True, color="FFFFFF")
    for index in range(1, columns + 1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"


def _autosize(sheet: Worksheet, section: RenderedSection) -> None:
    """Size columns from their content, sampling rows for speed."""
    cells = section.cell_values()
    for index, column in enumerate(section.columns, start=1):
        widest = len(column.label)
        for row in cells[:_SAMPLE_ROWS]:
            widest = max(widest, len(row[index - 1]))
        sheet.column_dimensions[get_column_letter(index)].width = min(widest + 2, _MAX_COLUMN_WIDTH)


def _write_table(sheet: Worksheet, section: RenderedSection, header_color: str) -> None:
    """Write one table section into its own worksheet.

    Values are written with their native Python types where possible so
    Excel sorts and aggregates them as numbers rather than as text --
    a spreadsheet where ``9`` sorts after ``10`` is worse than useless
    for the analysis these exports exist to support.
    """
    sheet.append([column.label for column in section.columns])
    for row in section.rows:
        sheet.append([_excel_value(row.get(column.key)) for column in section.columns])
    _style_header(sheet, len(section.columns), header_color)
    _autosize(sheet, section)
    if section.rows:
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(section.columns))}{len(section.rows) + 1}"
        )


def _excel_value(value: Any) -> Any:
    """Keep numbers/bools native; render everything else as text."""
    if value is None:
        return None
    if isinstance(value, bool | int | float):
        return value
    if isinstance(value, list | tuple):
        return ", ".join(str(item) for item in value)
    if isinstance(value, dict):
        return ", ".join(f"{key}={item}" for key, item in sorted(value.items()))
    return str(value)


def _add_chart(sheet: Worksheet, section: RenderedSection, start_row: int) -> None:
    """Write chart points and attach a native Excel chart.

    A real chart object rather than an image: it stays live, so a
    recipient can re-style it or extend the range, which is the entire
    reason to deliver a spreadsheet instead of a PDF.
    """
    sheet.cell(row=start_row, column=1, value="Label").font = Font(bold=True)
    sheet.cell(row=start_row, column=2, value="Value").font = Font(bold=True)
    for offset, (label, value) in enumerate(section.chart_points, start=1):
        sheet.cell(row=start_row + offset, column=1, value=label)
        sheet.cell(row=start_row + offset, column=2, value=value)

    last_row = start_row + len(section.chart_points)
    labels = Reference(sheet, min_col=1, min_row=start_row + 1, max_row=last_row)
    data = Reference(sheet, min_col=2, min_row=start_row, max_row=last_row)

    chart: BarChart | LineChart | PieChart
    if section.chart_kind is ChartKind.LINE:
        chart = LineChart()
    elif section.chart_kind is ChartKind.PIE:
        chart = PieChart()
    else:
        chart = BarChart()
    chart.title = section.title or section.key
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    sheet.add_chart(chart, f"D{start_row}")


def to_xlsx(report: RenderedReport) -> bytes:
    """Render *report* to XLSX bytes."""
    header_color, _stripe = theme_colors(report.branding.theme)
    workbook = Workbook()

    summary = workbook.active
    assert summary is not None
    summary.title = "Summary"
    summary.append(["Report", report.title])
    if report.subtitle:
        summary.append(["Subtitle", report.subtitle])
    summary.append(["Generated", report.generated_at.isoformat()])
    summary.append(["Company", report.branding.company_name])
    summary.append(["Total rows", report.total_rows])
    if report.parameters:
        summary.append([])
        summary.append(["Parameters", ""])
        for key, value in sorted(report.parameters.items()):
            summary.append([key, _excel_value(value)])
    for row in summary.iter_rows(min_col=1, max_col=1):
        row[0].font = Font(bold=True)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 60

    used_names = {"summary"}
    for section in report.sections:
        if section.error:
            summary.append([])
            summary.append([f"Section '{section.key}' unavailable", section.error])
            continue
        has_table = bool(section.rows and section.columns)
        if not has_table and not section.chart_points:
            if section.metric_value is not None:
                summary.append([section.metric_label or section.key, section.metric_value])
            continue

        sheet = workbook.create_sheet(_safe_sheet_name(section.title or section.key, used_names))
        if has_table:
            _write_table(sheet, section, header_color)
        if section.chart_points:
            _add_chart(sheet, section, start_row=(len(section.rows) + 3) if has_table else 1)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


__all__ = ["to_xlsx"]
