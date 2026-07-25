"""Excel (``.xlsx``) import/export parsing, via ``openpyxl``.

No ``shared_core`` equivalent exists -- confirmed by this service's own
reuse research, which also found the one existing precedent in this
monorepo (``services/user-management-service/app/parsers
/excel_parser.py``), reused here verbatim.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook, load_workbook


def parse_excel_rows(content: bytes) -> list[dict[str, str]]:
    """Parse *content* (an ``.xlsx`` file's raw bytes) into a list of header-keyed row dicts.

    The first row is treated as the header.
    """
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        return []
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = [str(cell) if cell is not None else "" for cell in next(rows_iter)]
    except StopIteration:
        return []
    rows: list[dict[str, str]] = []
    for values in rows_iter:
        row = {
            header[i]: ("" if value is None else str(value))
            for i, value in enumerate(values)
            if i < len(header)
        }
        if any(row.values()):
            rows.append(row)
    return rows


def write_excel_rows(rows: list[dict[str, Any]], fieldnames: list[str]) -> bytes:
    """Serialize *rows* to ``.xlsx`` bytes, with *fieldnames* as the column order."""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([row.get(field, "") for field in fieldnames])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


__all__ = ["parse_excel_rows", "write_excel_rows"]
